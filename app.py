import os
import math
from functools import lru_cache
from flask import Flask, jsonify, request, render_template, send_from_directory
import openpyxl
from google import genai
from google.genai import types as genai_types


import os
from dotenv import load_dotenv # Make sure this is installed: pip install python-dotenv
from google import genai



BASE_DIR = os.path.dirname(os.path.abspath(__file__))

WORKSPACE_ROOT = os.path.dirname(os.path.dirname(BASE_DIR))

DATA_PATH = os.path.join(BASE_DIR, "public_emdat_incl_hist_2026-05-18.xlsx") 
# Or adjust the filename to match exactly what you have

app = Flask(__name__, template_folder="templates", static_folder="static")

# ─── Data loading ────────────────────────────────────────────────────────────

_cached_data = None

def num(v):
    if v is None or v == "":
        return 0
    try:
        n = float(v)
        return 0 if math.isnan(n) else n
    except (TypeError, ValueError):
        return 0

def load_data():
    global _cached_data
    if _cached_data is not None:
        return _cached_data
    wb = openpyxl.load_workbook(DATA_PATH, read_only=False, data_only=True)
    ws = wb["EM-DAT Data"]
    headers = [str(cell) if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        year = num(r.get("Start Year"))
        if year and year > 0:
            rows.append(r)
    wb.close()
    _cached_data = rows
    return rows

def filter_data(data, start_year=None, end_year=None, disaster_type=None, region=None, country=None):
    out = []
    for r in data:
        y = num(r.get("Start Year"))
        if start_year and y < start_year:
            continue
        if end_year and y > end_year:
            continue
        if disaster_type and r.get("Disaster Type") != disaster_type:
            continue
        if region and r.get("Region") != region:
            continue
        if country and r.get("Country") != country:
            continue
        out.append(r)
    return out

def qint(v, default=None):
    try:
        return int(v) if v is not None else default
    except (ValueError, TypeError):
        return default

def qstr(v):
    return str(v) if v else None


_cached_summary = None

def build_data_summary():
    global _cached_summary
    if _cached_summary:
        return _cached_summary

    data = load_data()

    total_events = len(data)
    total_deaths = total_affected = total_damage = total_injured = total_homeless = 0

    year_map = {}
    decade_map = {}
    type_map = {}
    subtype_map = {}
    group_map = {}
    region_map = {}
    subregion_map = {}
    country_map = {}
    region_type_map = {}
    country_type_map = {}
    country_type_stats_map = {}

    def mk(): return {"count": 0, "deaths": 0, "affected": 0, "damage": 0, "injured": 0, "homeless": 0}
    def mk2(): return {"count": 0, "deaths": 0, "affected": 0, "damage": 0}
    def add(s, r):
        s["count"] += 1
        s["deaths"] += num(r.get("Total Deaths"))
        s["affected"] += num(r.get("Total Affected"))
        s["damage"] += num(r.get("Total Damage ('000 US$)"))
        s["injured"] += num(r.get("No. Injured"))
        s["homeless"] += num(r.get("No. Homeless"))

    for r in data:
        year = int(num(r.get("Start Year")))
        dtype = r.get("Disaster Type") or "Unknown"
        dstype = r.get("Disaster Subtype") or "Unknown"
        dgroup = r.get("Disaster Group") or "Unknown"
        region = r.get("Region") or "Unknown"
        subreg = r.get("Subregion") or "Unknown"
        country = r.get("Country") or "Unknown"
        decade = f"{(year // 10) * 10}s" if year else "Unknown"
        deaths = num(r.get("Total Deaths"))
        affected = num(r.get("Total Affected"))
        damage = num(r.get("Total Damage ('000 US$)"))

        total_deaths += deaths
        total_affected += affected
        total_damage += damage
        total_injured += num(r.get("No. Injured"))
        total_homeless += num(r.get("No. Homeless"))

        for m, k in [(year_map, year), (type_map, dtype), (subtype_map, dstype),
                     (group_map, dgroup), (region_map, region), (subregion_map, subreg),
                     (country_map, country), (decade_map, decade)]:
            if k not in m:
                m[k] = mk()
            add(m[k], r)

        region_type_map.setdefault(region, {})
        region_type_map[region][dtype] = region_type_map[region].get(dtype, 0) + 1

        country_type_map.setdefault(country, {})
        country_type_map[country][dtype] = country_type_map[country].get(dtype, 0) + 1

        # Full stats: country × disaster type
        country_type_stats_map.setdefault(country, {})
        if dtype not in country_type_stats_map[country]:
            country_type_stats_map[country][dtype] = mk2()
        s2 = country_type_stats_map[country][dtype]
        s2["count"] += 1
        s2["deaths"] += deaths
        s2["affected"] += affected
        s2["damage"] += damage

    years_sorted = sorted((k, v) for k, v in year_map.items() if isinstance(k, int))
    min_year = years_sorted[0][0] if years_sorted else 0
    max_year = years_sorted[-1][0] if years_sorted else 0

    year_by_year = " ".join(f"{y}:{s['count']}" for y, s in years_sorted)

    top20_events = "\n  ".join(
        f"{y}: {s['count']} events, {round(s['deaths']):,} deaths, {round(s['affected']):,} affected"
        for y, s in sorted(year_map.items(), key=lambda x: -x[1]['count'])[:20]
        if isinstance(y, int)
    )

    top15_deaths = "\n  ".join(
        f"{y}: {round(s['deaths']):,} deaths ({s['count']} events)"
        for y, s in sorted(year_map.items(), key=lambda x: -x[1]['deaths'])[:15]
        if isinstance(y, int)
    )

    top10_damage = ", ".join(
        f"{y}: ${s['damage']*1000/1e9:.1f}B"
        for y, s in sorted(((k, v) for k, v in year_map.items() if isinstance(k, int) and v['damage'] > 0), key=lambda x: -x[1]['damage'])[:10]
    )

    decade_breakdown = "\n  ".join(
        f"{d}: {s['count']} events, {round(s['deaths']):,} deaths, {round(s['affected']):,} affected, ${s['damage']*1000/1e9:.1f}B damage"
        for d, s in sorted(((k, v) for k, v in decade_map.items() if k != "Unknown"), key=lambda x: x[0])
    )

    type_breakdown = "\n  ".join(
        f"{t}: {s['count']} events, {round(s['deaths']):,} deaths, {round(s['affected']):,} affected, ${s['damage']*1000/1e9:.1f}B damage"
        for t, s in sorted(type_map.items(), key=lambda x: -x[1]['count'])
    )

    type_by_deaths = ", ".join(
        f"{t}: {round(s['deaths']):,} deaths"
        for t, s in sorted(type_map.items(), key=lambda x: -x[1]['deaths'])
    )

    type_by_damage = ", ".join(
        f"{t}: ${s['damage']*1000/1e9:.1f}B"
        for t, s in sorted(((k, v) for k, v in type_map.items() if v['damage'] > 0), key=lambda x: -x[1]['damage'])
    )

    subtype_breakdown = "\n  ".join(
        f"{t}: {s['count']} events, {round(s['deaths']):,} deaths"
        for t, s in sorted(((k, v) for k, v in subtype_map.items() if k != "Unknown"), key=lambda x: -x[1]['count'])[:30]
    )

    group_breakdown = ", ".join(
        f"{g}: {s['count']} events, {round(s['deaths']):,} deaths"
        for g, s in sorted(group_map.items(), key=lambda x: -x[1]['count'])
    )

    region_breakdown = "\n  ".join(
        f"{rg}: {s['count']} events, {round(s['deaths']):,} deaths, {round(s['affected']):,} affected, ${s['damage']*1000/1e9:.1f}B damage"
        for rg, s in sorted(region_map.items(), key=lambda x: -x[1]['count'])
    )

    subregion_breakdown = "\n  ".join(
        f"{sr}: {s['count']} events, {round(s['deaths']):,} deaths"
        for sr, s in sorted(((k, v) for k, v in subregion_map.items() if k != "Unknown"), key=lambda x: -x[1]['count'])
    )

    region_type_breakdown = "\n".join(
        f"  {rg}: " + ", ".join(f"{t}({c})" for t, c in sorted(tm.items(), key=lambda x: -x[1]))
        for rg, tm in sorted(region_type_map.items(), key=lambda x: -sum(x[1].values()))
    )

    countries_by_events = sorted(country_map.items(), key=lambda x: -x[1]['count'])

    top30_events = "\n  ".join(
        f"{c}: {s['count']} events, {round(s['deaths']):,} deaths, {s['affected']/1e6:.1f}M affected, ${s['damage']*1000/1e9:.1f}B damage"
        for c, s in countries_by_events[:30]
    )

    top20_deaths = ", ".join(
        f"{c}: {round(s['deaths']):,} deaths"
        for c, s in sorted(country_map.items(), key=lambda x: -x[1]['deaths'])[:20]
    )

    top15_damage = ", ".join(
        f"{c}: ${s['damage']*1000/1e9:.1f}B"
        for c, s in sorted(((k, v) for k, v in country_map.items() if v['damage'] > 0), key=lambda x: -x[1]['damage'])[:15]
    )

    country_type_breakdown = "\n".join(
        f"  {c}: " + ", ".join(f"{t}({n})" for t, n in sorted(country_type_map.get(c, {}).items(), key=lambda x: -x[1]))
        for c, _ in countries_by_events[:80]
        if c in country_type_map
    )

    # ── Top countries by deaths/affected/damage for EACH disaster type ──────────
    all_types = sorted(type_map.keys())
    def top_countries_for_type_by(dtype, metric, top_n=15):
        entries = [
            (c, country_type_stats_map[c][dtype])
            for c in country_type_stats_map
            if dtype in country_type_stats_map[c] and country_type_stats_map[c][dtype][metric] > 0
        ]
        entries.sort(key=lambda x: -x[1][metric])
        if metric == "damage":
            return ", ".join(f"{c}: ${s[metric]*1000/1e9:.2f}B" for c, s in entries[:top_n])
        elif metric == "affected":
            return ", ".join(f"{c}: {round(s[metric]):,}" for c, s in entries[:top_n])
        else:
            return ", ".join(f"{c}: {round(s[metric]):,} deaths" for c, s in entries[:top_n])

    per_type_country_deaths = "\n".join(
        f"  {dtype} — top by deaths: " + top_countries_for_type_by(dtype, "deaths")
        for dtype in all_types if dtype != "Unknown"
    )
    per_type_country_affected = "\n".join(
        f"  {dtype} — top by affected: " + top_countries_for_type_by(dtype, "affected")
        for dtype in all_types if dtype != "Unknown"
    )
    per_type_country_damage = "\n".join(
        f"  {dtype} — top by damage: " + top_countries_for_type_by(dtype, "damage")
        for dtype in all_types if dtype != "Unknown" and any(
            country_type_stats_map[c].get(dtype, {}).get("damage", 0) > 0
            for c in country_type_stats_map
        )
    )
    per_type_country_events = "\n".join(
        f"  {dtype} — top by events: " + ", ".join(
            f"{c}: {s['count']} events"
            for c, s in sorted(
                [(c, country_type_stats_map[c][dtype]) for c in country_type_stats_map if dtype in country_type_stats_map[c]],
                key=lambda x: -x[1]["count"]
            )[:15]
        )
        for dtype in all_types if dtype != "Unknown"
    )

    all_countries = ", ".join(c for c, _ in countries_by_events)

    _cached_summary = f"""EM-DAT GLOBAL DISASTER DATABASE — COMPLETE ANALYTICAL SUMMARY
Dataset: {total_events:,} events · {min_year}–{max_year} · {len(country_map)} countries

GLOBAL TOTALS:
  Total deaths: {round(total_deaths):,}
  Total affected: {round(total_affected):,}
  Total injured: {round(total_injured):,}
  Total homeless: {round(total_homeless):,}
  Total economic damage: ${total_damage*1000/1e12:.2f}T USD

━━━ YEAR-BY-YEAR EVENT COUNT ━━━
  {year_by_year}

━━━ TOP 20 YEARS BY EVENT COUNT ━━━
  {top20_events}

━━━ TOP 15 YEARS BY DEATHS ━━━
  {top15_deaths}

━━━ TOP 10 YEARS BY ECONOMIC DAMAGE ━━━
  {top10_damage}

━━━ DECADE BREAKDOWN ━━━
  {decade_breakdown}

━━━ DISASTER GROUP BREAKDOWN ━━━
  {group_breakdown}

━━━ DISASTER TYPE BREAKDOWN (all types) ━━━
  {type_breakdown}

━━━ DISASTER TYPES RANKED BY DEATHS ━━━
  {type_by_deaths}

━━━ DISASTER TYPES RANKED BY DAMAGE ━━━
  {type_by_damage}

━━━ TOP 30 DISASTER SUBTYPES ━━━
  {subtype_breakdown}

━━━ REGION BREAKDOWN ━━━
  {region_breakdown}

━━━ SUBREGION BREAKDOWN ━━━
  {subregion_breakdown}

━━━ DISASTER TYPES PER REGION ━━━
{region_type_breakdown}

━━━ TOP 30 COUNTRIES BY EVENT COUNT ━━━
  {top30_events}

━━━ TOP 20 COUNTRIES BY DEATHS ━━━
  {top20_deaths}

━━━ TOP 15 COUNTRIES BY ECONOMIC DAMAGE ━━━
  {top15_damage}

━━━ DISASTER TYPE BREAKDOWN PER COUNTRY (top 80, event counts) ━━━
{country_type_breakdown}

━━━ TOP COUNTRIES BY DEATHS FOR EACH DISASTER TYPE ━━━
{per_type_country_deaths}

━━━ TOP COUNTRIES BY PEOPLE AFFECTED FOR EACH DISASTER TYPE ━━━
{per_type_country_affected}

━━━ TOP COUNTRIES BY ECONOMIC DAMAGE FOR EACH DISASTER TYPE ━━━
{per_type_country_damage}

━━━ TOP COUNTRIES BY EVENT COUNT FOR EACH DISASTER TYPE ━━━
{per_type_country_events}

━━━ ALL {len(country_map)} COUNTRIES IN DATASET ━━━
  {all_countries}""".strip()

    return _cached_summary

# ─── Helpers ──────────────────────────────────────────────────────────────────

def compact(n):
    """Format big number as compact string (1.2M, 4.9T, etc.)"""
    n = abs(n)
    if n >= 1e12: return f"{n/1e12:.1f}T"
    if n >= 1e9:  return f"{n/1e9:.1f}B"
    if n >= 1e6:  return f"{n/1e6:.1f}M"
    if n >= 1e3:  return f"{n/1e3:.1f}K"
    return str(int(n))

# ─── Routes ───────────────────────────────────────────────────────────────────

BASE = ""

@app.route("/")
def index():
    return render_template("index.html")

@app.route(BASE + "/api/disasters/summary")
def summary():
    data = load_data()
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        disaster_type=qstr(request.args.get("disasterType")),
        region=qstr(request.args.get("region")),
    )
    total_events = len(filtered)
    total_deaths = sum(num(r.get("Total Deaths")) for r in filtered)
    total_affected = sum(num(r.get("Total Affected")) for r in filtered)
    total_damage_usd = sum(num(r.get("Total Damage ('000 US$)")) for r in filtered) * 1000
    years = [num(r.get("Start Year")) for r in filtered if num(r.get("Start Year")) > 0]
    min_year = int(min(years)) if years else 0
    max_year = int(max(years)) if years else 0
    countries = len(set(r.get("Country") for r in filtered if r.get("Country")))
    types = len(set(r.get("Disaster Type") for r in filtered if r.get("Disaster Type")))
    return jsonify({
        "totalEvents": total_events,
        "totalDeaths": round(total_deaths),
        "totalAffected": round(total_affected),
        "totalDamageUsd": total_damage_usd,
        "yearsSpan": f"{min_year}–{max_year}",
        "countriesAffected": countries,
        "disasterTypes": types,
    })

@app.route(BASE + "/api/disasters/by-year")
def by_year():
    data = load_data()
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        disaster_type=qstr(request.args.get("disasterType")),
        region=qstr(request.args.get("region")),
    )
    by_year_map = {}
    for r in filtered:
        y = int(num(r.get("Start Year")))
        if not y: continue
        if y not in by_year_map:
            by_year_map[y] = {"count": 0, "deaths": 0, "affected": 0, "damage": 0}
        by_year_map[y]["count"] += 1
        by_year_map[y]["deaths"] += num(r.get("Total Deaths"))
        by_year_map[y]["affected"] += num(r.get("Total Affected"))
        by_year_map[y]["damage"] += num(r.get("Total Damage ('000 US$)"))
    return jsonify([
        {"year": y, "count": v["count"], "deaths": round(v["deaths"]),
         "affected": round(v["affected"]), "damage": round(v["damage"])}
        for y, v in sorted(by_year_map.items())
    ])

@app.route(BASE + "/api/disasters/by-type")
def by_type():
    data = load_data()
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        region=qstr(request.args.get("region")),
    )
    counts = {}
    for r in filtered:
        t = r.get("Disaster Type") or "Unknown"
        counts[t] = counts.get(t, 0) + 1
    return jsonify([{"category": k, "count": v} for k, v in sorted(counts.items(), key=lambda x: -x[1])])

@app.route(BASE + "/api/disasters/by-region")
def by_region():
    data = load_data()
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        disaster_type=qstr(request.args.get("disasterType")),
    )
    counts = {}
    for r in filtered:
        rg = r.get("Region") or "Unknown"
        counts[rg] = counts.get(rg, 0) + 1
    return jsonify([{"category": k, "count": v} for k, v in sorted(counts.items(), key=lambda x: -x[1])])

@app.route(BASE + "/api/disasters/by-country")
def by_country():
    data = load_data()
    limit = qint(request.args.get("limit"), 15)
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        disaster_type=qstr(request.args.get("disasterType")),
    )
    m = {}
    for r in filtered:
        c = r.get("Country") or "Unknown"
        if c not in m:
            m[c] = {"iso": r.get("ISO") or "", "count": 0, "deaths": 0, "affected": 0}
        m[c]["count"] += 1
        m[c]["deaths"] += num(r.get("Total Deaths"))
        m[c]["affected"] += num(r.get("Total Affected"))
    result = sorted(m.items(), key=lambda x: -x[1]["count"])[:limit]
    return jsonify([{"country": c, "iso": v["iso"], "count": v["count"],
                     "deaths": round(v["deaths"]), "affected": round(v["affected"])} for c, v in result])

@app.route(BASE + "/api/disasters/deaths-by-type")
def deaths_by_type():
    data = load_data()
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        region=qstr(request.args.get("region")),
    )
    m = {}
    for r in filtered:
        t = r.get("Disaster Type") or "Unknown"
        m[t] = m.get(t, 0) + num(r.get("Total Deaths"))
    return jsonify([{"category": k, "value": round(v)} for k, v in sorted(m.items(), key=lambda x: -x[1])])

@app.route(BASE + "/api/disasters/damage-by-type")
def damage_by_type():
    data = load_data()
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        region=qstr(request.args.get("region")),
    )
    m = {}
    for r in filtered:
        t = r.get("Disaster Type") or "Unknown"
        m[t] = m.get(t, 0) + num(r.get("Total Damage ('000 US$)"))
    return jsonify([{"category": k, "value": round(v * 1000)}
                    for k, v in sorted(m.items(), key=lambda x: -x[1]) if v > 0])

@app.route(BASE + "/api/disasters/by-month")
def by_month():
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    data = load_data()
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        disaster_type=qstr(request.args.get("disasterType")),
    )
    counts = [0] * 12
    for r in filtered:
        m = int(num(r.get("Start Month") or 0))
        if 1 <= m <= 12:
            counts[m - 1] += 1
    return jsonify([{"month": i+1, "monthName": MONTHS[i], "count": counts[i]} for i in range(12)])

@app.route(BASE + "/api/disasters/affected-trend")
def affected_trend():
    data = load_data()
    filtered = filter_data(
        data,
        disaster_type=qstr(request.args.get("disasterType")),
        region=qstr(request.args.get("region")),
    )
    decades = {}
    for r in filtered:
        y = int(num(r.get("Start Year")))
        if not y: continue
        d = f"{(y // 10) * 10}s"
        if d not in decades:
            decades[d] = {"deaths": 0, "affected": 0, "events": 0}
        decades[d]["deaths"] += num(r.get("Total Deaths"))
        decades[d]["affected"] += num(r.get("Total Affected"))
        decades[d]["events"] += 1
    return jsonify([
        {"period": d, "deaths": round(v["deaths"]), "affected": round(v["affected"]), "events": v["events"]}
        for d, v in sorted(decades.items())
    ])

@app.route(BASE + "/api/disasters/aid-by-region")
def aid_by_region():
    data = load_data()
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
    )
    m = {}
    for r in filtered:
        rg = r.get("Region") or "Unknown"
        m[rg] = m.get(rg, 0) + num(r.get("AID Contribution ('000 US$)"))
    return jsonify([{"category": k, "value": round(v * 1000)}
                    for k, v in sorted(m.items(), key=lambda x: -x[1]) if v > 0])

@app.route(BASE + "/api/disasters/custom-chart")
def custom_chart():
    data = load_data()
    x_field = request.args.get("xField", "Disaster Type")
    y_field = request.args.get("yField", "Count")
    limit = qint(request.args.get("limit"), 20)
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        disaster_type=qstr(request.args.get("disasterType")),
        region=qstr(request.args.get("region")),
        country=qstr(request.args.get("country")),
    )

    def get_x(r):
        m = {
            "Disaster Type": lambda r: r.get("Disaster Type") or "Unknown",
            "Disaster Subtype": lambda r: r.get("Disaster Subtype") or "Unknown",
            "Region": lambda r: r.get("Region") or "Unknown",
            "Subregion": lambda r: r.get("Subregion") or "Unknown",
            "Country": lambda r: r.get("Country") or "Unknown",
            "Disaster Group": lambda r: r.get("Disaster Group") or "Unknown",
            "Decade": lambda r: f"{(int(num(r.get('Start Year'))) // 10) * 10}s",
        }
        return m.get(x_field, m["Disaster Type"])(r)

    def get_y(rows):
        m = {
            "Count": lambda rows: len(rows),
            "Total Deaths": lambda rows: sum(num(r.get("Total Deaths")) for r in rows),
            "Total Affected": lambda rows: sum(num(r.get("Total Affected")) for r in rows),
            "Total Damage (USD)": lambda rows: sum(num(r.get("Total Damage ('000 US$)")) * 1000 for r in rows),
            "AID Contribution (USD)": lambda rows: sum(num(r.get("AID Contribution ('000 US$)")) * 1000 for r in rows),
            "No. Injured": lambda rows: sum(num(r.get("No. Injured")) for r in rows),
            "No. Homeless": lambda rows: sum(num(r.get("No. Homeless")) for r in rows),
        }
        return m.get(y_field, m["Count"])(rows)

    grouped = {}
    for r in filtered:
        key = get_x(r)
        grouped.setdefault(key, []).append(r)

    chart_data = sorted(
        [{"label": k, "value": round(get_y(rows))} for k, rows in grouped.items()],
        key=lambda x: -x["value"]
    )[:limit]

    return jsonify({"data": chart_data, "xLabel": x_field, "yLabel": y_field})

@app.route(BASE + "/api/disasters/records")
def records():
    data = load_data()
    filtered = filter_data(
        data,
        start_year=qint(request.args.get("startYear")),
        end_year=qint(request.args.get("endYear")),
        disaster_type=qstr(request.args.get("disasterType")),
        region=qstr(request.args.get("region")),
        country=qstr(request.args.get("country")),
    )
    search = (request.args.get("search") or "").lower()
    if search:
        filtered = [r for r in filtered if
            search in (r.get("Country") or "").lower() or
            search in (r.get("Disaster Type") or "").lower() or
            search in (r.get("Disaster Subtype") or "").lower() or
            search in (r.get("Region") or "").lower() or
            search in (r.get("DisNo.") or "").lower()]

    sort_by = request.args.get("sortBy", "year")
    sort_dir = -1 if request.args.get("sortDir", "desc") == "desc" else 1

    sort_keys = {
        "year": lambda r: num(r.get("Start Year")),
        "type": lambda r: r.get("Disaster Type") or "",
        "country": lambda r: r.get("Country") or "",
        "region": lambda r: r.get("Region") or "",
        "deaths": lambda r: num(r.get("Total Deaths")),
        "affected": lambda r: num(r.get("Total Affected")),
        "damage": lambda r: num(r.get("Total Damage ('000 US$)")),
        "aid": lambda r: num(r.get("AID Contribution ('000 US$)")),
    }
    key_fn = sort_keys.get(sort_by, sort_keys["year"])
    reverse = sort_dir == -1
    filtered.sort(key=key_fn, reverse=reverse)

    page = max(1, qint(request.args.get("page"), 1))
    page_size = min(200, max(10, qint(request.args.get("pageSize"), 50)))
    total = len(filtered)
    start = (page - 1) * page_size
    paged = filtered[start:start + page_size]

    def row_out(r):
        dmg = num(r.get("Total Damage ('000 US$)"))
        aid = num(r.get("AID Contribution ('000 US$)"))
        return {
            "disNo": r.get("DisNo.") or "",
            "year": int(num(r.get("Start Year"))) or None,
            "month": int(num(r.get("Start Month"))) if r.get("Start Month") else None,
            "type": r.get("Disaster Type") or "",
            "subtype": r.get("Disaster Subtype") or "",
            "country": r.get("Country") or "",
            "iso": r.get("ISO") or "",
            "region": r.get("Region") or "",
            "subregion": r.get("Subregion") or "",
            "deaths": round(num(r.get("Total Deaths"))) or None,
            "injured": round(num(r.get("No. Injured"))) or None,
            "affected": round(num(r.get("Total Affected"))) or None,
            "homeless": round(num(r.get("No. Homeless"))) or None,
            "damage": round(dmg * 1000) if dmg > 0 else None,
            "aid": round(aid * 1000) if aid > 0 else None,
        }

    return jsonify({"data": [row_out(r) for r in paged], "total": total, "page": page, "pageSize": page_size})

@app.route(BASE + "/api/disasters/filters")
def filters():
    data = load_data()
    types = sorted(set(r.get("Disaster Type") for r in data if r.get("Disaster Type")))
    regions = sorted(set(r.get("Region") for r in data if r.get("Region")))
    countries = sorted(set(r.get("Country") for r in data if r.get("Country")))
    years = [num(r.get("Start Year")) for r in data if num(r.get("Start Year")) > 0]
    return jsonify({
        "disasterTypes": types,
        "regions": regions,
        "countries": countries,
        "minYear": int(min(years)) if years else 1900,
        "maxYear": int(max(years)) if years else 2026,
    })

@app.route(BASE + "/api/ai/chat", methods=["POST"])
def ai_chat():
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        return jsonify({"answer": "Gemini API key not configured."}), 500

    body = request.get_json() or {}
    message = body.get("message", "")
    history = body.get("history", [])

    if not message:
        return jsonify({"answer": "Message is required."}), 400

    data_summary = build_data_summary()

    client = genai.Client(api_key=api_key)

    system_prompt = f"""You are a concise data analyst for the EM-DAT Global Disaster Database.

Rules:
- Answer in 2–4 sentences max. Get straight to the number or finding — no preamble.
- State the key number first, then one sentence of context if useful.
- Plain prose only — no markdown, bullets, or headers.
- You have the COMPLETE dataset below: every year, every disaster type, every region, every country, subtypes, decades, and cross-tabulations. Use it to answer any question confidently with exact numbers.
- If a chart in the dashboard already visualises the answer, mention it briefly at the end.

Complete dataset summary:
{data_summary}"""

    contents = [
        genai_types.Content(role="user", parts=[genai_types.Part(text=system_prompt)]),
        genai_types.Content(role="model", parts=[genai_types.Part(text="Understood. I have the complete EM-DAT dataset — all years, disaster types, regions, countries, subtypes, and cross-breakdowns. Ask anything.")]),
    ]
    for m in history:
        role = "model" if m.get("role") == "assistant" else "user"
        contents.append(genai_types.Content(role=role, parts=[genai_types.Part(text=m.get("content", ""))]))
    contents.append(genai_types.Content(role="user", parts=[genai_types.Part(text=message)]))

    response = client.models.generate_content(model="gemini-2.5-flash", contents=contents)
    return jsonify({"answer": response.text})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    load_data()
    app.run(host="0.0.0.0", port=port, debug=False)
